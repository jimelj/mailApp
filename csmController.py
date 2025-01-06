import os
import math
import pandas as pd
from datetime import datetime
import zipfile
from tabulate import tabulate
import platform
from openpyxl import load_workbook
from PySide6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QLabel, QHeaderView, QPushButton, QFileDialog, QHBoxLayout
import webbrowser
from util import upload_to_ftps
from dotenv import load_dotenv
# Conditional import for Windows-specific packages
if platform.system() == "Windows":
    import win32com.client





def parse_zip_and_prepare_data(zip_file_path):
    """Extracts a ZIP file and parses the CSM data."""
    extracted_folder = "data/extracted"

    # Ensure the extracted folder exists
    os.makedirs(extracted_folder, exist_ok=True)

    # Extract the ZIP file
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(extracted_folder)

    # Locate the CSM file
    maildat_folder = os.path.join(extracted_folder, "MailDats")
    csm_files = [file for file in os.listdir(maildat_folder) if file.endswith(".csm")]

    if not csm_files:
        raise FileNotFoundError("No CSM file found in the uploaded ZIP file.")

    csm_file_path = os.path.join(maildat_folder, csm_files[0])
    return parse_csm_file(csm_file_path)


def parse_csm_file(csm_file_path):
    """Parses the CSM file and prepares the filtered DataFrame."""
    fields_positions = [
        ("Job ID", 1, 8),
        ("Segment ID", 9, 12),
        ("Container Type", 13, 14),
        ("Container ID", 15, 20),
        ("Display Container ID", 21, 26),
        ("Container Destination Zip", 36, 41),
        ("Container Level", 42, 43),
        ("Entry Point - Postal Code", 44, 49),
        ("Entry Point - Facility Type", 50, 51),
        ("Entry Point - Actual/Delivery Locale Key", 52, 60),
        ("Entry Point - Actual Postal Code", 61, 69),
        ("Parent Container Reference ID", 70, 75),
        ("Truck or Dispatch Number", 76, 95),
        ("Stop Designator", 96, 97),
        ("Reservation Number", 98, 112),
        ("Actual Container Ship Date", 113, 120),
        ("Actual Container Ship Time", 121, 125),
        ("Scheduled Pick Up Date", 126, 133),
        ("Scheduled Pick Up Time", 134, 138),
        ("Scheduled In-Home Date", 139, 146),
        ("Additional In-Home Range", 147, 147),
        ("Scheduled Induction Start Date", 148, 155),
        ("Scheduled Induction Start Time", 156, 160),
        ("Scheduled Induction End Date", 161, 168),
        ("Scheduled Induction End Time", 169, 173),
        ("Actual Induction Date", 174, 181),
        ("Actual Induction Time", 182, 186),
        ("Postage Statement Mailing Date", 187, 194),
        ("Postage Statement Mailing Time", 195, 199),
        ("Number of Copies", 200, 207),
        ("Number of Pieces", 208, 215),
        ("Total Weight", 216, 227),
        ("Container Status", 240, 240),
        ("Included in Other Documentation", 241, 241),
        ("Tray Preparation Type", 242, 242),
        ("Trans-Ship Bill of Lading Number", 243, 252),
        ("Sibling Container Indicator", 253, 253),
        ("Sibling Container Reference ID", 254, 259),
        ("Postage Grouping ID", 260, 267),
        ("Container Gross Weight", 268, 279),
        ("Container Height", 280, 282),
        ("EMD ASN Barcode", 283, 302),
        ("Transportation Carrier ID", 303, 317),
        ("FAST Content ID", 318, 326),
        ("FAST Scheduler ID", 327, 338),
        ("USPS Pick Up", 339, 339),
        ("CSA Separation ID", 340, 342),
        ("Scheduled Ship Date", 343, 350),
        ("Scheduled Ship Time", 351, 355),
        ("DMM Section Defining Container Preparation", 356, 367),
        ("Label: IM™ Container - Final", 368, 391),
        ("Label: IM™ Container - Original", 392, 415),
        ("Label: Destination Line 1", 416, 445),
        ("Label: Destination Line 2", 446, 475),
        ("Label: Contents - Line 1", 476, 505),
        ("Label: Contents - Line 2", 506, 525),
        ("Label: Entry Point Line", 526, 555),
        ("Label: User Information Line 1", 556, 595),
        ("Label: User Information Line 2", 596, 635),
        ("Label: Container Label CIN Code", 636, 639),
        ("eInduction Indicator", 660, 660),
        ("CSA Agreement ID", 661, 670),
        ("Presort Labeling List Effective Date", 671, 678),
        ("Last Used Labeling List Effective Date", 679, 686),
        ("Presort City-State Publication Date", 687, 694),
        ("Last Used City-State Publication Date", 695, 702),
        ("Presort Zone Chart Matrix Publication Date", 703, 710),
        ("Last Used Zone Chart Matrix Publication Date", 711, 718),
        ("Last Used Mail Direction Publication Date", 719, 726),
        ("Supplemental Physical Container ID", 727, 732),
        ("Accept Misshipped", 733, 733),
        ("Referenceable Mail Start Date", 734, 741),
        ("Referenceable Mail End Date", 742, 749),
        ("CSM Record Status", 750, 750),
        ("Reserve", 751, 789),
        ("Closing Character", 790, 790),
    ]

    def convert_weight(weight_field):
        """Converts weight to pounds and appends 'LBS'."""
        if weight_field:
            try:
                weight_in_lbs = math.ceil(float(f"{weight_field[:-4]}.{weight_field[-4:]}"))
                return f"{weight_in_lbs} LBS"
            except ValueError:
                return None
        return None

    def format_date(date_field):
        """Formats dates to MM-DD-YYYY."""
        try:
            if not date_field or len(date_field) != 8:
                return None
            return datetime.strptime(date_field, "%Y%m%d").strftime("%m-%d-%Y")
        except ValueError:
            return None

    # parsed_records = []
    # with open(csm_file_path, 'r') as file:
    #     for record in file.readlines():
    #         parsed_record = {}
    #         for field_name, start, end in fields_positions:
    #             raw_value = record[start - 1:end].strip()
    #             if field_name in {"Scheduled Induction Start Date"}:
    #                 raw_value = format_date(raw_value)
    #             elif field_name == "Total Weight":
    #                 raw_value = convert_weight(raw_value)
    #             parsed_record[field_name] = raw_value
    #         parsed_records.append(parsed_record)

# Parse all records from the CSM file
    parsed_records = []
    with open(csm_file_path, 'r') as file:
        for record in file.readlines():
            parsed_record = {field[0]: record[field[1]-1:field[2]].strip() for field in fields_positions if record[field[1]-1:field[2]].strip()}

        # Convert Total Weight to pounds
            if "Total Weight" in parsed_record:
                parsed_record["Total Weight"] = convert_weight(parsed_record["Total Weight"])
            
            # Format Scheduled Induction Start Date
            if "Scheduled Induction Start Date" in parsed_record:
                parsed_record["Scheduled Induction Start Date"] = format_date(parsed_record["Scheduled Induction Start Date"])
            
            # Ensure Number of Pieces is represented properly
            if "Number of Pieces" in parsed_record:
                parsed_record["Number of Pieces"] = int(parsed_record["Number of Pieces"]) if parsed_record["Number of Pieces"].isdigit() else None
            
        


            parsed_records.append(parsed_record)


    df_csm = pd.DataFrame(parsed_records)

    # Save full parsed data
    df_csm.to_csv("data/parsed_csm.csv", index=False)
    # print("Full parsed data saved to 'data/parsed_csm.csv'")

    # Print full parsed data to terminal
    # print("\nFull Parsed Data (Terminal):")
    # print(tabulate(df_csm, headers="keys", tablefmt="pretty", showindex=False))

    def match_facility_address(parsed_csm_path, facility_report_path):
        """
        Matches the last 6 characters of 'Entry Point - Actual/Delivery Locale Key'
        in the parsed CSM file with 'Dropsite Key' in the facility report and 
        returns the formatted address.

        Parameters:
            parsed_csm_path (str): Path to the parsed CSM CSV file.
            facility_report_path (str): Path to the facility report Excel file.

        Returns:
            list: A list of formatted address strings for matches.
        """
        try:
            # Load the parsed CSM and facility report data
            parsed_csm = pd.read_csv(parsed_csm_path)
            facility_report = pd.read_excel(facility_report_path, header=1)

            # Extract the last 6 characters for comparison
            parsed_csm['Last_6_Locale_Key'] = parsed_csm['Entry Point - Actual/Delivery Locale Key'].str[-6:]
            facility_report['Last_6_Dropsite_Key'] = facility_report['Dropsite Key'].astype(str).str[-6:]

            # Aggregate facility data if needed (e.g., ensure unique keys)
            facility_report_aggregated = facility_report.groupby('Last_6_Dropsite_Key').first().reset_index()

            print("Before Merge: Parsed CSM")
            print(parsed_csm.head())

            # Perform the merge to add information from facility_report to parsed_csm
            parsed_csm = pd.merge(
                parsed_csm,
                facility_report_aggregated[['Last_6_Dropsite_Key', 'Address', 'City', 'State', 'ZIP Code']],
                left_on='Last_6_Locale_Key',
                right_on='Last_6_Dropsite_Key',
                how='left'
            )

            print("After Merge: Parsed CSM")
            print(parsed_csm.head())

            # Format the matched address
            parsed_csm['ZIP Code'] = parsed_csm['ZIP Code'].astype(str)  # Ensure ZIP Code is a string
            parsed_csm['Address'] = parsed_csm.apply(
                lambda row: f"{row['Address']}, {row['City']}, {row['State']}, {row['ZIP Code'][:5]}"
                if pd.notnull(row['Address']) and pd.notnull(row['ZIP Code'])
                else None,
                axis=1
            )

            # Drop unnecessary columns added during the merge
            parsed_csm.drop(columns=['Last_6_Locale_Key', 'Last_6_Dropsite_Key'], inplace=True)

            print("Hello WORLD")
            print(parsed_csm)

            return parsed_csm
        # try:
        #     # Load the parsed CSM and facility report data
        #     parsed_csm = pd.read_csv(parsed_csm_path)
        #     facility_report = pd.read_excel(facility_report_path, header=1)
            
        #     # Extract the last 6 characters for comparison
        #     parsed_csm['Last_6_Locale_Key'] = parsed_csm['Entry Point - Actual/Delivery Locale Key'].str[-6:]
        #     facility_report['Last_6_Dropsite_Key'] = facility_report['Dropsite Key'].astype(str).str[-6:]
        #     # parsed_csm['Last_6_Locale_Key'] = parsed_csm['Entry Point - Actual/Delivery Locale Key'].astype(str).str[-6:]
        #     # facility_report['Last_6_Dropsite_Key'] = facility_report['Dropsite Key'].astype(str).str[-6:]

        #     # Debugging: Print the extracted keys
        #     print("Parsed CSM Last 6 Characters:")
        #     print(parsed_csm[['Entry Point - Actual/Delivery Locale Key', 'Last_6_Locale_Key']].head())
        #     print("\nFacility Report Last 6 Characters:")
        #     print(facility_report[['Dropsite Key', 'Last_6_Dropsite_Key']].head())
            
        #     # Perform the comparison
        #     matches = pd.merge(
        #         parsed_csm[['Last_6_Locale_Key']],
        #         facility_report[['Last_6_Dropsite_Key', 'Address', 'City', 'State', 'ZIP Code']],
        #         left_on='Last_6_Locale_Key',
        #         right_on='Last_6_Dropsite_Key',
        #         how='inner'
        #     )
            
        #     # Format the matched address as 'Address, City, State, ZIP'
        #     # Ensure ZIP Code is a string and format the matched addresses
        #     # matches['ZIP Code'] = matches['ZIP Code'].astype(str)  # Ensure ZIP Code is a string
        #     # formatted_addresses = matches.apply(
        #     #     lambda row: f"{row['Address']}, {row['City']}, {row['State']}, {row['ZIP Code'][:5]}",
        #     #     axis=1
        #     # ).tolist()

        #     # Create the 'Matched Address' column
        #     matches['ZIP Code'] = matches['ZIP Code'].astype(str)  # Ensure ZIP Code is a string
        #     matches['Address'] = matches.apply(
        #     lambda row: f"{row['Address']}, {row['City']}, {row['State']}, {row['ZIP Code'][:5]}"
        #     if pd.notnull(row['Address']) else None,
        #     axis=1
        # )
            
        #      # Merge the addresses back into the parsed CSM data
        #     parsed_csm = pd.merge(
        #     parsed_csm,
        #     matches[['Last_6_Locale_Key', 'Address']],
        #     on='Last_6_Locale_Key',
        #     how='left'
        # )
        #     print('I AM AROUND HERE')
        #     print(parsed_csm)
        #     return parsed_csm
            
        except Exception as e:
            return f"An error occurred: {str(e)}"
        
    if __name__ == "__main__":
        # Example: Testing the new function
        # parsed_csm_path = "data/parsed_csm.csv"  # Replace with actual path
        # facility_report_path = "facilityReport.xlsx"  # Replace with actual path
        parsed_csm_path = os.path.join("data", "parsed_csm.csv") #crossplatform
        facility_report_path = os.path.join("facilityReport.xlsx") #crossplatform
        print ("Hello world")
        results = match_facility_address(parsed_csm_path, facility_report_path)
        
        if isinstance(results, list):
            print("Matching Facility Addresses:")
            for address in results:
                print(address)
        else:
            print(results)

    
    print("I AM HERE")
    UpdatedParsedCSM = match_facility_address("data/parsed_csm.csv", "facilityReport.xlsx")
    print(UpdatedParsedCSM)

    # Filtered fields for app display
    app_display_fields = [
        "Job ID",
        "Display Container ID",
        "Container Destination Zip",
        "Label: Destination Line 1",
        "Scheduled Induction Start Date",
        "Number of Pieces",
        "Total Weight",
        "Label: IM™ Container - Final",
        "Address",
    ]
    df_filtered = UpdatedParsedCSM[app_display_fields]

    # Print filtered data to terminal
    print("\nFiltered Data for App Display:")
    print(tabulate(df_filtered, headers="keys", tablefmt="pretty", showindex=False))

    return df_filtered


class CSMTab(QWidget):
    """CSM Tab for displaying parsed CSM data."""
    def __init__(self, df_filtered):
        super().__init__()
        self.df_filtered = df_filtered
        self.processed_zip_name = None
        self.layout = QVBoxLayout(self)

        self.table = QTableWidget()
        self.layout.addWidget(self.table)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()

        # "Email Report" button
        self.email_button = QPushButton("Email CMS Report")
        self.email_button.clicked.connect(self.email_report)  # Connect button to functionality

        # "Capstone" button
        self.capstone_button = QPushButton("Capstone Report")
        self.capstone_button.clicked.connect(self.generate_capstone_report)


        # Button Styling
        button_style = """
            QPushButton {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 16px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """
        self.email_button.setStyleSheet(button_style)
        self.capstone_button.setStyleSheet(button_style)
        # Add buttons to the horizontal layout
        button_layout.addWidget(self.email_button)
        button_layout.addWidget(self.capstone_button)

        # Add button layout to the main vertical layout
        self.layout.addLayout(button_layout)
        

        self.update_table()
    
    def set_processed_zip_name(self, processed_zip_name):
        """
        Set the processed ZIP name to be used later.
        """
        self.processed_zip_name = processed_zip_name
        print(f"Processed ZIP Name set in CSMTab: {self.processed_zip_name}")


    def update_data(self, df_filtered):
        """Updates the tab with new data."""
        self.df_filtered = df_filtered
        self.update_table()

    # def update_table(self):
    #     """Updates the table widget to display the DataFrame."""
    #     if self.df_filtered.empty:
    #         self.table.clear()
    #         self.table.setRowCount(0)
    #         self.table.setColumnCount(0)
    #         self.layout.addWidget(QLabel("No CSM data available. Please upload a ZIP file."))
    #     else:
    #         self.table.setRowCount(len(self.df_filtered))
    #         self.table.setColumnCount(len(self.df_filtered.columns))
    #         self.table.setHorizontalHeaderLabels(self.df_filtered.columns)

    #         for row_idx, row in self.df_filtered.iterrows():
    #             for col_idx, value in enumerate(row):
    #                 self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))

    #         header = self.table.horizontalHeader()
    #         for col in range(len(self.df_filtered.columns)):
    #             header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
    def update_table(self):
        """Updates the table widget to display the DataFrame."""
        # Check if a "No CSM data available" QLabel exists, and remove it if present
        for i in reversed(range(self.layout.count())):
            widget = self.layout.itemAt(i).widget()
            if isinstance(widget, QLabel) and "No CSM data available" in widget.text():
                widget.deleteLater()  # Safely delete the label widget

        if self.df_filtered.empty:
            print("DataFrame is empty!")  # Debugging
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            if not any(isinstance(self.layout.itemAt(i).widget(), QLabel) for i in range(self.layout.count())):
                self.layout.addWidget(QLabel("No CSM data available. Please upload a ZIP file."))
        else:
            # print("DataFrame is not empty!")  # Debugging
            self.table.setRowCount(len(self.df_filtered))
            self.table.setColumnCount(len(self.df_filtered.columns))
            self.table.setHorizontalHeaderLabels(self.df_filtered.columns)

            for row_idx, row in self.df_filtered.iterrows():
                for col_idx, value in enumerate(row):
                    self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))

            header = self.table.horizontalHeader()
            for col in range(len(self.df_filtered.columns)):
                header.setSectionResizeMode(col, QHeaderView.ResizeToContents)

            # Ensure the table is visible in the layout
            if self.table not in [self.layout.itemAt(i).widget() for i in range(self.layout.count())]:
                self.layout.addWidget(self.table)

    def email_report(self):
            """Generates an Excel report and opens the default email system."""
            if self.df_filtered.empty:
                print("No data available to generate the report.")
                return

            # Save the filtered DataFrame to an Excel file
            csm_report_name = f"CSM_Report {self.processed_zip_name}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(self, "Save Report", csm_report_name, "Excel Files (*.xlsx)")
            if file_path:
                try:
                    self.df_filtered.to_excel(file_path, index=False)

                     # Adjust column widths using openpyxl
                    workbook = load_workbook(file_path)
                    sheet = workbook.active

                    for column_cells in sheet.columns:
                        max_length = 0
                        column_letter = column_cells[0].column_letter  # Get the column letter (e.g., "A")
                        for cell in column_cells:
                            try:
                                if cell.value:  # Check if the cell has a value
                                    max_length = max(max_length, len(str(cell.value)))
                            except Exception as e:
                                print(f"Error calculating column width: {e}")
                        adjusted_width = max_length + 2  # Add padding
                        sheet.column_dimensions[column_letter].width = adjusted_width

                    workbook.save(file_path)
                    print('filepath')
                    print(file_path)

                    # # Open default email client with the file attached
                    # subject = "CSM Report"
                    # body = "Please find the attached CSM report."
                    # email_uri = f"mailto:?subject={subject}&body={body}&attachment={file_path}"
                    # print(f"Opening email client with URI: {email_uri}")
                    # webbrowser.open(email_uri)
                                # Create the email command
                    if platform.system() == "Windows":
                        # Create an Outlook application object
                        outlook = win32com.client.Dispatch("Outlook.Application")

                        # Create a new mail item
                        mail = outlook.CreateItem(0)

                        # Set the recipient, subject, and body
                        mail.To = "jjoseph@cbaol.com; deasterwood@cbaol.com;"
                        mail.Subject = "CSM Report"
                        mail.Body = "Attached is the CSM Report."

                        # Attach a file
                        mail.Attachments.Add(file_path)

                        # Send the email
                        mail.Send()
                    elif platform.system() == "Darwin":  # macOS
                        os.system(f"open 'mailto:?subject=CSM Report&body=Attached is the CSM Report&attachment={file_path}'")
                    else:  # Linux (may vary based on desktop environment)
                        webbrowser.open(f"mailto:?subject=CSM Report&body=Attached is the CSM Report&attachment={file_path}")
                except Exception as e:
                    print(f"Error creating or sending the report: {e}")
    
    def generate_capstone_report(self):
        """Generate the Capstone Excel report with specified fields."""
        load_dotenv()
        if self.df_filtered.empty:
            print("No data available to generate the Capstone report.")
            return

        try:
            # Split the Address column
            self.df_filtered[["Destination Address", "Destination City", "Destination State", "Destination Zip"]] = (
                self.df_filtered["Address"]
                .str.extract(r"^(.*?),\s*(.*?),\s*([A-Z]{2}),\s*(\d{5})$")
        )
            # Create the Capstone data structure with placeholders
            capstone_data = {
                "Customer Number*": ["11769"] * len(self.df_filtered), 
                "Billing Group": "",
                "Origin Name*": ["CBA Industries"] * len(self.df_filtered),
                "Origin Address*": ["160 Raritan Center Parkway"]* len(self.df_filtered),
                "Origin Suite": ["Suite 19"] * len(self.df_filtered), 
                "Origin City*": ["Edison"] * len(self.df_filtered), 
                "Origin State*": ["NJ"] * len(self.df_filtered), 
                "Origin Zip*": ["08837"] * len(self.df_filtered),
                "Origin Plus 4": "", 
                "Origin Phone": ["404-579-4090"] * len(self.df_filtered),  # Placeholder
                "Origin Remarks": ["Contact Darrell Easterwood"] * len(self.df_filtered),  # Placeholder
                "Destination Name*": self.df_filtered["Label: Destination Line 1"],  # Placeholder
                "Destination Address*": self.df_filtered["Destination Address"],
                "Destination Suite": "",
                "Destination City*": self.df_filtered["Destination City"],  # Placeholder
                "Destination State*": self.df_filtered["Destination State"],  # Placeholder
                "Destination Zip*": self.df_filtered["Destination Zip"],
                "Destination Plus 4": "",  # Placeholder
                "Destination Phone": "",
                "Destination Remarks": "",
                "Email Address": ["deasterwood@cbaol.com"] * len(self.df_filtered),
                "Send Confirmation Email": "",
                "Send POP Email": "",
                "Send POD Email": "",
                "Reference 1": self.df_filtered["Job ID"],
                "Reference 2": self.df_filtered["Display Container ID"],
                "Order Type*": "",  # Placeholder
                "Pieces": self.df_filtered["Number of Pieces"],
                "Weight": self.df_filtered["Total Weight"].str.replace(" LBS", "", regex=False),
                "Pickup Date*": self.df_filtered["Scheduled Induction Start Date"] + " 3:00 AM",
                "Driver ID": "",
                "Order Comments": "",
                "Parcel Barcode": self.df_filtered["Label: IM™ Container - Final"],
                "Parcel Pieces": "",
                "Parcel Length": "",
                "Parcel Width": "",
                "Parcel Height": "",
                "Parcel Weight": "",
            }

            # Convert to DataFrame
            capstone_df = pd.DataFrame(capstone_data)

            # Save the Capstone report to a CSV file
            capstone_report_name = f"Capstone_Report {self.processed_zip_name}.CSV"
            file_path, _ = QFileDialog.getSaveFileName(self, "Save Capstone Report", capstone_report_name, "CSV Files (*.csv)")
            if file_path:
                capstone_df.to_csv(file_path, index=False)

                # # Adjust column widths using openpyxl
                # workbook = load_workbook(file_path)
                # sheet = workbook.active

                # for column_cells in sheet.columns:
                #     max_length = 0
                #     column_letter = column_cells[0].column_letter  # Get the column letter (e.g., "A")
                #     for cell in column_cells:
                #         try:
                #             if cell.value:  # Check if the cell has a value
                #                 max_length = max(max_length, len(str(cell.value)))
                #         except Exception as e:
                #             print(f"Error calculating column width: {e}")
                #     adjusted_width = max_length + 2  # Add padding
                #     sheet.column_dimensions[column_letter].width = adjusted_width

                # workbook.save(file_path)
                # print(f"Capstone report saved to {file_path}")

                # host = str(os.getenv("HOSTNAME"))
                # username = str(os.getenv("USERNAME"))
                # password = str(os.getenv("FTP_SECRET"))
                # port = int(os.getenv("PORT"))
                # remote_dir = "/Capstone"

                
                host = os.getenv("HOSTNAME")
                username = os.getenv("FTP_USERNAME")
                password = os.getenv("FTP_SECRET")
                remote_dir = os.getenv("REMOTEDIR")
                port= int(os.getenv("PORT", 990))

                print('Rocket Fuel')
                print(username)

                result = upload_to_ftps(file_path, host, username, password, remote_dir, port)
            #     result = upload_to_ftps(file_path, 
            #    "cba.sharefileftp.com", 
            #    "cba/jjoseph@cbaol.com", 
            #    "JimelJoseph@15951", 
            #    "/Capstone", 
            #    990)
                print(result)

               
                    

        except Exception as e:
            print(f"Error creating Capstone report: {e}")